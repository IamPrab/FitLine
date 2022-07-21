import csv
import re
from typing import NamedTuple
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import json

class Results:
    TFM_Count: int
    LFM_Count: int
    HFM_Count: int
    Unknown_Count: int
    TFM_Kill_Count: int
    LFM_Kill_Count: int
    HFM_Kill_Count: int
    Unknown_Kill_Count: int
    TFM_Escape_Count: int
    LFM_Escape_Count: int
    HFM_Escape_Count: int
    Unknown_Escape_Count: int

class Dye(NamedTuple):  # Unit Datastructures to store cells from database
    lot: str
    wafer: int
    bin: int
    X: int
    Y: int
    testname: str
    resultstring: str
    modulename: str
    status: str
    fmcategory: str
    delta: int
    offset: int


def post_testname(testname):  # extract information from testname
    result = []
    step_0 = testname.split('::')

    modulename = step_0[0]

    result.append(modulename)
    if len(step_0)>=2:
        status = "KILL"  # Default
        fmcategory = "unk"
        s = step_0[1]
        if s.find("_EDC") > 0:  # Need to test
            status = "EDC"  # Check if you have both what category would that be?
        elif s.find("_E_") > 0:
            status = "EDC"

        if s.find("LFM") > 0:
            fmcategory = "LFM"
        elif s.find("TFM") > 0:
            fmcategory = "TFM"
        elif s.find("HFM") > 0:
            fmcategory = "HFM"

        result.append(status)
        result.append(fmcategory)
    else:
        result.append(testname)
        result.append("unk")
        result.append("unk")
    # print(result)
    return result


def post_resultstring(name):  #
    values = name.split('|')
    del_off = []
    delta=0
    offset = 0  # default# VMIN_OFFSET_004 VMIN_EDC??, UNDEFINED??
    if (len(values)==8):
        X = float(values[0])
        Y = float(values[1])
        intercept = float(values[2])
        slope = float(values[3])
        # print(values[7])
        if "SIGMA" in values[7]:
            if re.findall(r'\d+', values[7]):
                num = re.findall(r'\d+', values[7])
                offset = float(num[0])
        elif "OP" in values[7]:
            if re.findall(r'\d+', values[7]):
                num = re.findall(r'\d+', values[7])
                offset = float(num[0])/2
            # print(offset/0.02)
        elif "VMIN" in values[7]:
            if re.findall(r'\d+', values[7]):
                num =re.findall(r'\d+', values[7])
                offset = float(num[0]) * 5
        elif "0." in values[7]:
            if re.findall(r'\d+', values[7]):
                num = re.findall(r'\d+', values[7])
                offset = float(num[0])/2

        delta = round(((slope * X) + intercept - Y) / 0.02)

    del_off = [delta, offset]
    return del_off


def map_dyes(dye_single, blocks):  # blocks with modulenames

    key = dye_single.modulename +"%"+ dye_single.status + "%" + dye_single.fmcategory
    if key in blocks:
        blocks[key].append(dye_single)
    else:
        blocks[key] = [dye_single]

    return blocks



def read_data(file):
    with open(file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        blocks = {}
        for row in csv_reader:
            if line_count > 0:
                r = post_testname(row[8])

                del_off = post_resultstring(row[9])
                Dye_single = Dye(row[0], row[3], int(row[6]), row[4], row[5], row[8],
                                 row[9],
                                 r[0], r[1], r[2], del_off[0], del_off[1])
                blocks = map_dyes(Dye_single, blocks)
            line_count += 1
        # print(len(blocks["SCN_GRT"]))
    return blocks



def cout_FmModuleStatus(fmCat):
    a={}
    for key in fmCat:
        for dye in fmCat[key]:
            key1 = dye.status + "%" + dye.testname + "%" + dye.modulename +"%" + dye.fmcategory
            if key1 in fmCat:
                a[key1]= a[key1]+1
            else:
                a[key1] = 1

    fm_module_cat={}
    for key in a:
        step_0 = key.split("%")
        status= step_0[0]
        module= step_0[2]
        fmcategory=step_0[3]

        key1= module +"%" + status +"%" + fmcategory

        if key1 in fm_module_cat:
            fm_module_cat[key1]=fm_module_cat[key1]+1
        else:
            fm_module_cat[key1]=1
    #print(fm_module_cat)

    return fm_module_cat


def count_escapes(fmCat):
    escape = {}

    for key in fmCat:
        escape[key] = 0
        for dye in fmCat[key]:
            if dye.delta < 0:
                escape[key] = escape[key] + 1


    #print(escape)
    return escape

def red_alert(fmCat, file):
    reds = {}
    name= "Escapes"+ file
    with open(name, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["LOT", "WAFER", "X", "Y", "MODULE NAME", "TEST_NAME", "BIN", "DELTA", "STATUS"])
        for key in fmCat:
            reds[key] = 0
            for dye in fmCat[key]:
                if dye.delta < 0 and dye.bin < 3 and dye.status =="KILL":
                    reds[key] = reds[key] + 1
                    writer.writerow(
                        [dye.lot, dye.wafer, dye.X, dye.Y, dye.modulename, dye.testname, dye.bin, dye.delta, dye.status])

    #print(reds)
    return reds


def result_file_1(writer,fm_module_Cat,sheet,escapes,reds):
    module_cat = {}
    writer.writerow(["Module Name","STATUS", "N(Total)", "N(TFM)", "N(LFM)", "N(HFM)", "N(UNK)"," "
                      "NK(Total)", "NK(TFM)","E(TFM)", "NK(LFM)","E(LFM)", "NK(HFM)","E(HFM)", "NK(UNK)","E(UNK)"])

    for key in fm_module_Cat:
        name = key.split("%")
        module = name[0]
        status = name[1]
        fmcat= name[2]
        key1= module +"@" +status
        if key1 in module_cat:
            if fmcat == 'LFM':
                lfm= fm_module_Cat[key]
            elif fmcat == 'TFM':
                tfm= fm_module_Cat[key]
            elif fmcat == 'HFM':
                hfm= fm_module_Cat[key]
            elif fmcat == 'unk':
                unk= fm_module_Cat[key]
        else:
            lfm = 0
            tfm = 0
            hfm = 0
            unk = 0
            if fmcat == 'LFM':
                lfm= fm_module_Cat[key]
            elif fmcat == 'TFM':
                tfm= fm_module_Cat[key]
            elif fmcat == 'HFM':
                hfm= fm_module_Cat[key]
            elif fmcat == 'unk':
                unk= fm_module_Cat[key]
        module_cat[key1]=[lfm,tfm,hfm,unk]
    #print(module_cat)

    module_cat2 = {}
    for key in escapes:
        name = key.split("%")
        module = name[0]
        status = name[1]
        fmcat= name[2]
        key1= module +"@" +status
        if key1 in module_cat2:
            if fmcat == 'LFM':
                lfm= escapes[key]
                lfm_red= reds[key]
            elif fmcat == 'TFM':
                tfm= escapes[key]
                tfm_red = reds[key]
            elif fmcat == 'HFM':
                hfm= escapes[key]
                hfm_red = reds[key]
            elif fmcat == 'unk':
                unk= escapes[key]
                unk_red = reds[key]
        else:
            lfm = 0
            tfm = 0
            hfm = 0
            unk = 0
            lfm_red=0
            tfm_red=0
            hfm_red=0
            unk_red=0
            if fmcat == 'LFM':
                lfm = escapes[key]
                lfm_red = reds[key]
            elif fmcat == 'TFM':
                tfm = escapes[key]
                tfm_red = reds[key]
            elif fmcat == 'HFM':
                hfm = escapes[key]
                hfm_red = reds[key]
            elif fmcat == 'unk':
                unk = escapes[key]
                unk_red = reds[key]
        module_cat2[key1]=[lfm,lfm_red,tfm,tfm_red,hfm,hfm_red,unk,unk_red]
    #print(module_cat)

    for key in module_cat2:
        name= key.split("@")
        result1= Results()
        result1.TFM_Count= module_cat[key][0]
        result1.LFM_Count = module_cat[key][1]
        result1.HFM_Count = module_cat[key][2]
        result1.Unknown_Count_Count = module_cat[key][3]

        result1.TFM_Kill_Count= module_cat2[key][0]
        result1.LFM_Kill_Count = module_cat2[key][2]
        result1.HFM_Kill_Count = module_cat2[key][4]
        result1.Unknown_Kill_Count = module_cat2[key][6]

        result1.TFM_Escape_Count = module_cat2[key][1]
        result1.LFM_Escape_Count = module_cat2[key][3]
        result1.HFM_Escape_Count = module_cat2[key][5]
        result1.Unknown_Escape_Count= module_cat2[key][7]

        jsonStr= json.dumps(result1.__dict__)
        print(jsonStr)

        writer.writerow([name[0],name[1],module_cat[key][0]+module_cat[key][1]+module_cat[key][2]+module_cat[key][3],
                         module_cat[key][0],module_cat[key][1],module_cat[key][2],module_cat[key][3]," ",
                         module_cat2[key][0]+module_cat2[key][2]+module_cat2[key][4]+module_cat2[key][6],module_cat2[key][0],module_cat2[key][1],
                         module_cat2[key][2],module_cat2[key][3],module_cat2[key][4],module_cat2[key][5],module_cat2[key][6],module_cat2[key][7]])
        row1=('','',name[0],name[1],module_cat[key][0]+module_cat[key][1]+module_cat[key][2]+module_cat[key][3],
                         module_cat[key][0],module_cat[key][1],module_cat[key][2],module_cat[key][3]," ",
                         module_cat2[key][0]+module_cat2[key][2]+module_cat2[key][4]+module_cat2[key][6],module_cat2[key][0],module_cat2[key][1],
                         module_cat2[key][2],module_cat2[key][3],module_cat2[key][4],module_cat2[key][5],module_cat2[key][6],module_cat2[key][7])
        sheet.append(row1)
        maxRow = sheet.max_row
        sheet.cell(row=maxRow, column=13).fill = PatternFill(fgColor='00FF0000', fill_type='solid')
        sheet.cell(row=maxRow, column=15).fill = PatternFill(fgColor='00FF0000', fill_type='solid')
        sheet.cell(row=maxRow, column=17).fill = PatternFill(fgColor='00FF0000', fill_type='solid')
        sheet.cell(row=maxRow, column=19).fill = PatternFill(fgColor='00FF0000', fill_type='solid')



if __name__ == '__main__':
    path= "C:/Users/kaurp/PycharmProjects/ADTL/Data1/NewTest.csv"
    if (os.path.exists('test.xlsx')):
        wb = load_workbook('test.xlsx')
        sheet = wb.active
    else:
        filepath = "C:/Users/kaurp/PycharmProjects/ADTL/test.xlsx"
        wb = Workbook()
        sheet = wb.active
    f = "res.csv"
    blocks = read_data(path)
    fm_module_cat = cout_FmModuleStatus(blocks)
    escapes = count_escapes(blocks)
    reds = red_alert(blocks,f)
    name= "summary"+f
    with open(name, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["MODULE Vs No. OF TESTS"])
        fill = PatternFill("solid", fgColor="00FF6600")
        fill2= PatternFill("solid", fgColor="00CCFFFF")
        row_to_fill_texts = ('testname', 'programename')
        names_to_fill_texts = ("", "", "Module Name", "STATUS", "N(Total)", "N(TFM)", "N(LFM)", "N(HFM)", "N(UNK)","", "NK(Total)"
                     ,"NK(TFM)", "E(TFM)", "NK(LFM)", "E(LFM)", "NK(HFM)", "E(HFM)", "NK(UNK)", "E(UNK)")
        row = []
        names=[]
        for h in row_to_fill_texts:
            cell = WriteOnlyCell(sheet, value=h)
            cell.fill = fill
            row.append(cell)
        for h in names_to_fill_texts:
            cell = WriteOnlyCell(sheet, value=h)
            cell.fill = fill2
            names.append(cell)
        sheet.append(row)
        sheet.append(names)
        result_file_1(writer, fm_module_cat,sheet, escapes, reds)
        wb.save("test.xlsx")



